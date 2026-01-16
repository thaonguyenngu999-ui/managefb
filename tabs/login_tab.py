"""
Tab Login FB - Đăng nhập Facebook cho các profiles
"""
import customtkinter as ctk
from typing import List, Dict, Optional
import threading
import time
import os
import queue
from datetime import datetime
from config import COLORS
from widgets import ModernCard, ModernButton, ModernEntry
from api_service import api

# Optional: openpyxl for XLSX support
try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None
    Workbook = None


class LoginTab(ctk.CTkFrame):
    """Tab đăng nhập Facebook cho profiles"""

    def __init__(self, master, status_callback=None, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)

        self.status_callback = status_callback
        self.folders: List[Dict] = []
        self.profiles: List[Dict] = []
        self.accounts: List[Dict] = []  # Từ file XLSX
        self.xlsx_path: str = ""
        self.workbook: Optional[Workbook] = None

        # Profile status: {uuid: {'has_fb': bool, 'status': str}}
        self.profile_status: Dict[str, Dict] = {}
        self.profile_vars: Dict[str, ctk.BooleanVar] = {}

        # Running state
        self._is_running = False
        self._stop_requested = False
        self._threads: List[threading.Thread] = []

        self._create_ui()
        self._load_folders()

    def _create_ui(self):
        """Tạo giao diện"""
        # ========== HEADER ==========
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 15))

        ctk.CTkLabel(
            header_frame,
            text="🔐 Đăng nhập Facebook",
            font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")

        ModernButton(
            header_frame,
            text="Làm mới",
            icon="🔄",
            variant="secondary",
            command=self._load_folders,
            width=100
        ).pack(side="right", padx=5)

        # ========== MAIN CONTENT - 2 COLUMNS ==========
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Left panel - Profile selection
        left_panel = ctk.CTkFrame(main_frame, fg_color=COLORS["bg_secondary"], corner_radius=15, width=400)
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        left_panel.pack_propagate(False)

        self._create_profile_panel(left_panel)

        # Right panel - Account import & Login
        right_panel = ctk.CTkFrame(main_frame, fg_color=COLORS["bg_secondary"], corner_radius=15)
        right_panel.pack(side="right", fill="both", expand=True)

        self._create_login_panel(right_panel)

    def _create_profile_panel(self, parent):
        """Tạo panel chọn profiles"""
        # Header
        header = ctk.CTkFrame(parent, fg_color="transparent")
        header.pack(fill="x", padx=15, pady=15)

        ctk.CTkLabel(
            header,
            text="📁 Profiles",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")

        # Folder selection
        folder_frame = ctk.CTkFrame(parent, fg_color="transparent")
        folder_frame.pack(fill="x", padx=15, pady=5)

        ctk.CTkLabel(folder_frame, text="Thư mục:", width=70, anchor="w").pack(side="left")
        self.folder_var = ctk.StringVar(value="-- Chọn --")
        self.folder_menu = ctk.CTkOptionMenu(
            folder_frame,
            variable=self.folder_var,
            values=["-- Chọn --"],
            fg_color=COLORS["bg_card"],
            button_color=COLORS["accent"],
            width=150,
            command=self._on_folder_change
        )
        self.folder_menu.pack(side="left", padx=5)

        ModernButton(
            folder_frame,
            text="Tải",
            variant="secondary",
            command=self._load_profiles,
            width=60
        ).pack(side="left", padx=5)

        # Check FB status controls
        check_frame = ctk.CTkFrame(parent, fg_color="transparent")
        check_frame.pack(fill="x", padx=15, pady=5)

        ctk.CTkLabel(check_frame, text="Luồng:", width=50, anchor="w").pack(side="left")
        self.check_thread_entry = ModernEntry(check_frame, placeholder="3", width=50)
        self.check_thread_entry.pack(side="left", padx=5)
        self.check_thread_entry.insert(0, "3")

        ModernButton(
            check_frame,
            text="🔍 Check FB",
            variant="primary",
            command=self._check_fb_status,
            width=100
        ).pack(side="left", padx=5)

        self.check_progress = ctk.CTkLabel(
            check_frame,
            text="",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        )
        self.check_progress.pack(side="right")

        # Filter controls
        filter_frame = ctk.CTkFrame(parent, fg_color="transparent")
        filter_frame.pack(fill="x", padx=15, pady=5)

        self.filter_var = ctk.StringVar(value="Tất cả")
        ctk.CTkRadioButton(
            filter_frame, text="Tất cả", variable=self.filter_var, value="Tất cả",
            command=self._apply_filter
        ).pack(side="left", padx=5)
        ctk.CTkRadioButton(
            filter_frame, text="Chưa có FB", variable=self.filter_var, value="Chưa có FB",
            command=self._apply_filter
        ).pack(side="left", padx=5)
        ctk.CTkRadioButton(
            filter_frame, text="Có FB", variable=self.filter_var, value="Có FB",
            command=self._apply_filter
        ).pack(side="left", padx=5)

        # Select all
        select_frame = ctk.CTkFrame(parent, fg_color="transparent")
        select_frame.pack(fill="x", padx=15, pady=5)

        self.select_all_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            select_frame,
            text="Chọn tất cả",
            variable=self.select_all_var,
            fg_color=COLORS["accent"],
            command=self._toggle_all
        ).pack(side="left")

        self.profile_count_label = ctk.CTkLabel(
            select_frame,
            text="(0 đã chọn)",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_secondary"]
        )
        self.profile_count_label.pack(side="left", padx=10)

        # Profile list
        self.profile_scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        self.profile_scroll.pack(fill="both", expand=True, padx=10, pady=(5, 10))

        ctk.CTkLabel(
            self.profile_scroll,
            text="Chọn thư mục để xem profiles",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_secondary"]
        ).pack(pady=30)

    def _create_login_panel(self, parent):
        """Tạo panel đăng nhập"""
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=15, pady=15)

        # Import XLSX section
        import_label = ctk.CTkLabel(
            scroll,
            text="📥 Import tài khoản Facebook",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=COLORS["text_primary"]
        )
        import_label.pack(anchor="w", pady=(0, 10))

        ctk.CTkLabel(
            scroll,
            text="XLSX: A(Status-trống), B(UID), C(Password), D(2FA Secret), E(Email), F(Email Pass)",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"],
            wraplength=400
        ).pack(anchor="w")

        import_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        import_frame.pack(fill="x", pady=10)

        self.xlsx_entry = ModernEntry(import_frame, placeholder="Chọn file XLSX...", width=300)
        self.xlsx_entry.pack(side="left", fill="x", expand=True)

        ModernButton(
            import_frame,
            text="📂 Chọn file",
            variant="secondary",
            command=self._browse_xlsx,
            width=100
        ).pack(side="left", padx=5)

        # Account preview
        self.account_info = ctk.CTkLabel(
            scroll,
            text="Chưa import file",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_secondary"]
        )
        self.account_info.pack(anchor="w", pady=5)

        # Account list preview
        self.account_list_frame = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10, height=150)
        self.account_list_frame.pack(fill="x", pady=10)
        self.account_list_frame.pack_propagate(False)

        self.account_scroll = ctk.CTkScrollableFrame(self.account_list_frame, fg_color="transparent")
        self.account_scroll.pack(fill="both", expand=True, padx=5, pady=5)

        ctk.CTkLabel(
            self.account_scroll,
            text="Import file XLSX để xem danh sách tài khoản",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(pady=30)

        # Login settings
        settings_label = ctk.CTkLabel(
            scroll,
            text="⚙️ Cài đặt Login",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=COLORS["text_primary"]
        )
        settings_label.pack(anchor="w", pady=(20, 10))

        settings_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        settings_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(settings_frame, text="Số luồng:", width=70, anchor="w").pack(side="left")
        self.login_thread_entry = ModernEntry(settings_frame, placeholder="3", width=60)
        self.login_thread_entry.pack(side="left", padx=5)
        self.login_thread_entry.insert(0, "3")

        ctk.CTkLabel(settings_frame, text="Delay (s):", width=70, anchor="w").pack(side="left", padx=(20, 0))
        self.delay_entry = ModernEntry(settings_frame, placeholder="5", width=60)
        self.delay_entry.pack(side="left", padx=5)
        self.delay_entry.insert(0, "5")

        # Options
        options_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        options_frame.pack(fill="x", pady=5)

        self.delete_bad_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            options_frame,
            text="🗑️ Xóa profile nếu nick DIE",
            variable=self.delete_bad_var,
            fg_color=COLORS["accent"]
        ).pack(side="left")

        self.save_xlsx_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            options_frame,
            text="Lưu trạng thái vào XLSX",
            variable=self.save_xlsx_var,
            fg_color=COLORS["accent"]
        ).pack(side="left", padx=20)

        # Folder đích khi login thành công
        dest_folder_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        dest_folder_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(dest_folder_frame, text="📁 Folder đích (LIVE):", width=130, anchor="w").pack(side="left")
        self.dest_folder_var = ctk.StringVar(value="-- Không chuyển --")
        self.dest_folder_menu = ctk.CTkOptionMenu(
            dest_folder_frame,
            variable=self.dest_folder_var,
            values=["-- Không chuyển --"],
            width=200,
            fg_color=COLORS["bg_card"],
            button_color=COLORS["accent"],
            dropdown_fg_color=COLORS["bg_secondary"]
        )
        self.dest_folder_menu.pack(side="left", padx=5)

        # Action buttons
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(20, 10))

        self.start_btn = ModernButton(
            btn_frame,
            text="▶️ Bắt đầu Login",
            variant="success",
            command=self._start_login,
            width=150
        )
        self.start_btn.pack(side="left", padx=5)

        self.stop_btn = ModernButton(
            btn_frame,
            text="⏹️ Dừng",
            variant="danger",
            command=self._stop_login,
            width=100
        )
        self.stop_btn.pack(side="left", padx=5)

        # Progress
        self.progress_label = ctk.CTkLabel(
            scroll,
            text="",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["accent"]
        )
        self.progress_label.pack(anchor="w", pady=5)

        self.progress_bar = ctk.CTkProgressBar(scroll, width=400, height=15)
        self.progress_bar.pack(fill="x", pady=5)
        self.progress_bar.set(0)

        # Log
        log_label = ctk.CTkLabel(
            scroll,
            text="📋 Log",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        )
        log_label.pack(anchor="w", pady=(15, 5))

        self.log_text = ctk.CTkTextbox(scroll, height=200, fg_color=COLORS["bg_card"])
        self.log_text.pack(fill="x", pady=5)

    def _load_folders(self):
        """Load danh sách thư mục"""
        try:
            self.folders = api.get_folders() or []
            folder_names = ["-- Chọn --"] + [f.get('name', '') for f in self.folders if f.get('name')]
            self.folder_menu.configure(values=folder_names)
            self.folder_var.set("-- Chọn --")

            # Cập nhật dropdown folder đích
            dest_folder_names = ["-- Không chuyển --"] + [f.get('name', '') for f in self.folders if f.get('name')]
            self.dest_folder_menu.configure(values=dest_folder_names)
        except Exception as e:
            self._log(f"Lỗi tải folders: {e}")

    def _get_dest_folder_uuid(self):
        """Lấy uuid của folder đích đã chọn"""
        try:
            dest_name = self.dest_folder_var.get()
            if dest_name == "-- Không chuyển --":
                return None

            for f in self.folders:
                if f.get('name', '').strip() == dest_name:
                    return f.get('uuid') or f.get('id')
        except Exception as e:
            print(f"Error getting dest folder: {e}")
        return None

    def _on_folder_change(self, choice):
        """Khi chọn folder"""
        if choice != "-- Chọn --":
            self._load_profiles()

    def _load_profiles(self):
        """Load profiles theo folder"""
        folder_name = self.folder_var.get()
        if folder_name == "-- Chọn --":
            return

        # Clear
        for widget in self.profile_scroll.winfo_children():
            widget.destroy()
        self.profile_vars = {}

        try:
            folder_id = None
            for f in self.folders:
                if f.get('name') == folder_name:
                    folder_id = f.get('id')
                    break

            if folder_id:
                self.profiles = api.get_profiles(folder_id=[folder_id], limit=500) or []
            else:
                self.profiles = []

            self._render_profiles()
            self._log(f"Đã tải {len(self.profiles)} profiles")

        except Exception as e:
            self._log(f"Lỗi tải profiles: {e}")

    def _render_profiles(self, filter_type: str = "Tất cả"):
        """Render danh sách profiles"""
        for widget in self.profile_scroll.winfo_children():
            widget.destroy()

        old_selections = {uuid: var.get() for uuid, var in self.profile_vars.items()}
        self.profile_vars = {}

        if not self.profiles:
            ctk.CTkLabel(
                self.profile_scroll,
                text="Không có profile nào",
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_secondary"]
            ).pack(pady=30)
            return

        for profile in self.profiles:
            uuid = profile.get('uuid', '')
            name = profile.get('name', uuid[:8])

            # Check filter
            status = self.profile_status.get(uuid, {})
            has_fb = status.get('has_fb', None)

            if filter_type == "Chưa có FB" and has_fb is True:
                continue
            if filter_type == "Có FB" and has_fb is not True:
                continue

            # Create frame for each profile
            pf = ctk.CTkFrame(self.profile_scroll, fg_color="transparent")
            pf.pack(fill="x", pady=2)

            var = ctk.BooleanVar(value=old_selections.get(uuid, False))
            self.profile_vars[uuid] = var

            cb = ctk.CTkCheckBox(
                pf,
                text=name,
                variable=var,
                fg_color=COLORS["accent"],
                command=self._update_count,
                width=150
            )
            cb.pack(side="left")

            # Status indicator
            if has_fb is True:
                status_text = "✅ Có FB"
                status_color = COLORS["success"]
            elif has_fb is False:
                status_text = "❌ Chưa có"
                status_color = COLORS["danger"]
            else:
                status_text = "❓ Chưa check"
                status_color = COLORS["text_secondary"]

            ctk.CTkLabel(
                pf,
                text=status_text,
                font=ctk.CTkFont(size=11),
                text_color=status_color
            ).pack(side="right", padx=5)

        self._update_count()

    def _apply_filter(self):
        """Apply profile filter"""
        self._render_profiles(self.filter_var.get())

    def _toggle_all(self):
        """Toggle select all"""
        select_all = self.select_all_var.get()
        for var in self.profile_vars.values():
            var.set(select_all)
        self._update_count()

    def _update_count(self):
        """Update selected count"""
        count = sum(1 for var in self.profile_vars.values() if var.get())
        self.profile_count_label.configure(text=f"({count} đã chọn)")

    def _check_fb_status(self):
        """Check if profiles have FB logged in"""
        if self._is_running:
            return

        selected = [uuid for uuid, var in self.profile_vars.items() if var.get()]
        if not selected:
            self._log("⚠️ Chưa chọn profile nào")
            return

        self._is_running = True
        self._stop_requested = False

        thread_count = int(self.check_thread_entry.get() or 3)

        def check_worker(profile_queue: queue.Queue):
            while not self._stop_requested:
                try:
                    uuid = profile_queue.get_nowait()
                except queue.Empty:
                    break

                try:
                    # Mở browser và check xem có FB không
                    result = api.open_browser(uuid)
                    if result.get('status') == 'successfully':
                        data = result.get('data', {})
                        remote_port = data.get('remote_port')
                        ws_url = data.get('web_socket', '')

                        # Kết nối CDP và check
                        has_fb = self._check_profile_has_fb(remote_port, ws_url)
                        self.profile_status[uuid] = {'has_fb': has_fb}

                        # Đóng browser
                        api.close_browser(uuid)

                        self.after(0, lambda u=uuid, h=has_fb: self._log(
                            f"[{u[:8]}] {'✅ Có FB' if h else '❌ Chưa có FB'}"
                        ))

                except Exception as e:
                    self.after(0, lambda u=uuid, err=str(e): self._log(f"[{u[:8]}] Lỗi: {err}"))

                profile_queue.task_done()

        def run_check():
            q = queue.Queue()
            for uuid in selected:
                q.put(uuid)

            threads = []
            for _ in range(min(thread_count, len(selected))):
                t = threading.Thread(target=check_worker, args=(q,))
                t.start()
                threads.append(t)

            for t in threads:
                t.join()

            self.after(0, self._on_check_complete)

        threading.Thread(target=run_check, daemon=True).start()

    def _check_profile_has_fb(self, remote_port: int, ws_url: str) -> bool:
        """Check if browser has FB logged in"""
        from automation import CDPHelper
        import time

        helper = CDPHelper()
        time.sleep(1.5)

        if not helper.connect(remote_port=remote_port, ws_url=ws_url):
            return False

        try:
            # Navigate to Facebook
            helper.navigate("https://www.facebook.com")
            helper.wait_for_page_ready(timeout_ms=10000)

            # Check if logged in by looking for profile elements
            js = '''
                (function() {
                    // Check for login form or logged-in indicators
                    let loginForm = document.querySelector('input[name="email"], input[name="pass"]');
                    if (loginForm) return false;

                    // Check for profile menu or user indicator
                    let profileMenu = document.querySelector('[aria-label*="Account"], [aria-label*="Tài khoản"]');
                    let messenger = document.querySelector('[aria-label*="Messenger"]');
                    let notifications = document.querySelector('[aria-label*="Notifications"], [aria-label*="Thông báo"]');

                    return !!(profileMenu || messenger || notifications);
                })()
            '''
            result = helper.execute_js(js)
            return result is True

        finally:
            helper.close()

    def _on_check_complete(self):
        """Khi check xong"""
        self._is_running = False
        self._apply_filter()
        self._log("✅ Đã check xong tất cả profiles")

    def _browse_xlsx(self):
        """Chọn file XLSX"""
        if not HAS_OPENPYXL:
            self._log("⚠️ Cần cài openpyxl: pip install openpyxl")
            return

        from tkinter import filedialog
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.xlsx_path = path
            self.xlsx_entry.delete(0, "end")
            self.xlsx_entry.insert(0, path)
            self._load_xlsx()

    def _load_xlsx(self):
        """Load dữ liệu từ file XLSX"""
        if not HAS_OPENPYXL:
            self._log("⚠️ Cần cài openpyxl: pip install openpyxl")
            return

        if not self.xlsx_path or not os.path.exists(self.xlsx_path):
            return

        try:
            self.workbook = openpyxl.load_workbook(self.xlsx_path)
            sheet = self.workbook.active

            self.accounts = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                if not row or len(row) < 3:  # Cần ít nhất B(UID) và C(Password)
                    continue

                # Columns: A=Status(trống), B=UID, C=Password, D=2FA Secret, E=Email, F=Email Pass
                # Helper: convert Excel float to clean string (61551817090446.0 -> "61551817090446")
                def clean_str(val):
                    if val is None:
                        return ''
                    if isinstance(val, float):
                        # Remove .0 from float numbers
                        return str(int(val)) if val == int(val) else str(val)
                    return str(val)

                account = {
                    'row': row_idx,
                    'status': clean_str(row[0]),
                    'fb_id': clean_str(row[1]),
                    'password': clean_str(row[2]),
                    'totp_secret': clean_str(row[3]) if len(row) > 3 else '',  # 2FA secret
                    'email': clean_str(row[4]) if len(row) > 4 else '',
                    'email_pass': clean_str(row[5]) if len(row) > 5 else ''
                }

                # Skip if no FB ID
                if account['fb_id']:
                    self.accounts.append(account)

            self.account_info.configure(text=f"Đã tải {len(self.accounts)} tài khoản")
            self._render_accounts()
            self._log(f"📥 Đã import {len(self.accounts)} tài khoản từ file")

        except Exception as e:
            self._log(f"Lỗi đọc file XLSX: {e}")

    def _render_accounts(self):
        """Render danh sách tài khoản"""
        for widget in self.account_scroll.winfo_children():
            widget.destroy()

        if not self.accounts:
            ctk.CTkLabel(
                self.account_scroll,
                text="Không có tài khoản nào",
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_secondary"]
            ).pack(pady=30)
            return

        for acc in self.accounts[:50]:  # Show first 50
            status = acc.get('status', '')
            if status == 'LIVE':
                color = COLORS["success"]
            elif status in ['DIE', 'ERROR']:
                color = COLORS["danger"]
            else:
                color = COLORS["text_secondary"]

            text = f"[{status or '?'}] {acc['fb_id'][:15]}... - {acc['username'][:20]}..."
            ctk.CTkLabel(
                self.account_scroll,
                text=text,
                font=ctk.CTkFont(size=11),
                text_color=color
            ).pack(anchor="w", pady=1)

        if len(self.accounts) > 50:
            ctk.CTkLabel(
                self.account_scroll,
                text=f"... và {len(self.accounts) - 50} tài khoản khác",
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_secondary"]
            ).pack(anchor="w", pady=5)

    def _update_xlsx_status(self, row: int, status: str):
        """Cập nhật trạng thái vào file XLSX"""
        if not self.workbook or not self.save_xlsx_var.get():
            return

        try:
            sheet = self.workbook.active
            sheet.cell(row=row, column=1, value=status)
            self.workbook.save(self.xlsx_path)
        except Exception as e:
            print(f"Error saving XLSX: {e}")

    def _start_login(self):
        """Bắt đầu đăng nhập"""
        if self._is_running:
            return

        selected_profiles = [uuid for uuid, var in self.profile_vars.items() if var.get()]
        if not selected_profiles:
            self._log("⚠️ Chưa chọn profile nào")
            return

        # Get accounts without status (not yet processed)
        available_accounts = [a for a in self.accounts if not a.get('status')]
        if not available_accounts:
            self._log("⚠️ Không có tài khoản nào chưa được xử lý")
            return

        self._is_running = True
        self._stop_requested = False

        thread_count = int(self.login_thread_entry.get() or 3)
        delay = int(self.delay_entry.get() or 5)

        self._log(f"▶️ Bắt đầu login {len(selected_profiles)} profiles với {len(available_accounts)} tài khoản")

        def login_worker(profile_queue: queue.Queue, account_queue: queue.Queue):
            while not self._stop_requested:
                try:
                    uuid = profile_queue.get_nowait()
                except queue.Empty:
                    break

                # Get profile info
                profile = next((p for p in self.profiles if p.get('uuid') == uuid), None)
                if not profile:
                    profile_queue.task_done()
                    continue

                profile_name = profile.get('name', uuid[:8])

                # Get an account
                try:
                    account = account_queue.get_nowait()
                except queue.Empty:
                    self.after(0, lambda pn=profile_name: self._log(f"[{pn}] Hết tài khoản"))
                    profile_queue.task_done()
                    break

                self.after(0, lambda pn=profile_name, fb=account['fb_id'][:10]:
                    self._log(f"[{pn}] Đang login với {fb}..."))

                try:
                    success, status = self._login_profile(uuid, account)

                    # Update account status
                    account['status'] = status
                    self._update_xlsx_status(account['row'], status)

                    if success:
                        self.after(0, lambda pn=profile_name: self._log(f"[{pn}] ✅ Login thành công"))
                        self.profile_status[uuid] = {'has_fb': True}

                        # Move profile vào folder đã chọn
                        try:
                            dest_folder = self._get_dest_folder_uuid()
                            dest_name = self.dest_folder_var.get()
                            if dest_folder:
                                move_result = api.add_profiles_to_folder(dest_folder, [uuid])
                                self.after(0, lambda dn=dest_name: self._log(f"[{profile_name}] 📁 Moved to {dn}"))
                        except Exception as e:
                            self.after(0, lambda err=str(e): self._log(f"  Move folder error: {err}"))
                    else:
                        self.after(0, lambda pn=profile_name, s=status: self._log(f"[{pn}] ❌ {s}"))
                        # Không retry - mỗi profile chỉ login 1 account

                except Exception as e:
                    self.after(0, lambda pn=profile_name, err=str(e): self._log(f"[{pn}] Lỗi: {err}"))

                time.sleep(delay)
                profile_queue.task_done()

        def run_login():
            profile_q = queue.Queue()
            account_q = queue.Queue()

            for uuid in selected_profiles:
                profile_q.put(uuid)
            for acc in available_accounts:
                account_q.put(acc)

            threads = []
            for _ in range(min(thread_count, len(selected_profiles))):
                t = threading.Thread(target=login_worker, args=(profile_q, account_q))
                t.start()
                threads.append(t)

            for t in threads:
                t.join()

            self.after(0, self._on_login_complete)

        threading.Thread(target=run_login, daemon=True).start()

    def _login_profile(self, uuid: str, account: Dict) -> tuple:
        """Login FB cho profile - dùng WebSocket trực tiếp (nhanh hơn)"""
        import websocket
        import json as json_module
        import requests

        result = api.open_browser(uuid)
        if result.get('status') != 'successfully':
            return False, 'ERROR'

        data = result.get('data', {})
        remote_port = data.get('remote_port')

        if not remote_port:
            return False, 'NO_PORT'

        ws = None
        try:
            # Lấy page WebSocket từ /json endpoint
            time.sleep(1)
            resp = requests.get(f"http://127.0.0.1:{remote_port}/json", timeout=5)
            tabs = resp.json()

            page_ws = None
            for tab in tabs:
                if tab.get('type') == 'page':
                    page_ws = tab.get('webSocketDebuggerUrl')
                    break

            if not page_ws:
                api.close_browser(uuid)
                return False, 'NO_PAGE'

            # Kết nối WebSocket trực tiếp (nhanh hơn CDPHelper)
            ws = websocket.create_connection(page_ws, timeout=15, suppress_origin=True)

            # Helper functions
            msg_id = [0]
            def send_cmd(method, params=None):
                msg_id[0] += 1
                msg = {"id": msg_id[0], "method": method}
                if params:
                    msg["params"] = params
                ws.send(json_module.dumps(msg))
                return json_module.loads(ws.recv())

            def evaluate(expression):
                result = send_cmd("Runtime.evaluate", {
                    "expression": expression,
                    "returnByValue": True
                })
                return result.get('result', {}).get('result', {}).get('value')

            # Xóa cookies Facebook trước khi login (tránh dính session cũ)
            for domain in [".facebook.com", "facebook.com", "www.facebook.com", "m.facebook.com"]:
                send_cmd("Network.deleteCookies", {"domain": domain})
            # Xóa storage
            for origin in ["https://www.facebook.com", "https://m.facebook.com"]:
                send_cmd("Storage.clearDataForOrigin", {"origin": origin, "storageTypes": "all"})

            # Navigate to Facebook login
            send_cmd("Page.navigate", {"url": "https://www.facebook.com/login"})
            time.sleep(3)  # Đợi page load

            # Check và click "Không phải bạn?" nếu có session cũ
            not_you_result = evaluate('''
                (function() {
                    // Tìm link/button "Không phải bạn?" hoặc "Not you?"
                    let els = document.querySelectorAll('a, div[role="button"], span, button');
                    for (let el of els) {
                        let text = (el.innerText || el.textContent || '').trim();
                        if (text.includes('Không phải bạn') || text.includes('Not you') ||
                            text.includes('Quên mật khẩu') === false && text.includes('khác') ||
                            text.includes('Log in with another') || text.includes('Đăng nhập bằng tài khoản khác')) {
                            el.click();
                            return 'clicked_not_you';
                        }
                    }
                    // Tìm theo aria-label
                    let notYou = document.querySelector('a[href*="login/identify"]') ||
                                document.querySelector('div[aria-label*="Không phải"]');
                    if (notYou) {
                        notYou.click();
                        return 'clicked_not_you_v2';
                    }
                    return 'no_session';
                })()
            ''')
            if not_you_result and 'clicked' in str(not_you_result):
                self.after(0, lambda: self._log(f"  Clicked 'Không phải bạn?'"))
                time.sleep(2)  # Đợi page reload

            import random

            # Helper: Simulate human typing với CDP Input.insertText
            def type_text(text, field_selector):
                """Gõ từng ký tự như người thật"""
                # Focus vào field trước
                evaluate(f'''
                    (function() {{
                        let el = document.querySelector('{field_selector}');
                        if (el) {{
                            el.focus();
                            el.value = '';  // Clear existing
                        }}
                    }})()
                ''')
                time.sleep(random.uniform(0.2, 0.5))

                # Gõ từng ký tự với Input.insertText (chuẩn CDP)
                for char in text:
                    send_cmd("Input.insertText", {"text": char})
                    # Random delay giữa các phím (100-280ms như người gõ thật)
                    time.sleep(random.uniform(0.10, 0.28))

            # Helper: Generate TOTP code từ secret
            def generate_totp(secret):
                """Generate 6-digit TOTP code"""
                import hmac
                import hashlib
                import struct

                # Clean secret (remove spaces, uppercase)
                secret = secret.replace(' ', '').upper()

                # Base32 decode
                base32_chars = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ234567'
                bits = ''
                for c in secret:
                    if c in base32_chars:
                        bits += bin(base32_chars.index(c))[2:].zfill(5)
                key = bytes(int(bits[i:i+8], 2) for i in range(0, len(bits) - len(bits) % 8, 8))

                # TOTP với time step 30s
                counter = int(time.time()) // 30
                counter_bytes = struct.pack('>Q', counter)

                # HMAC-SHA1
                hmac_hash = hmac.new(key, counter_bytes, hashlib.sha1).digest()

                # Dynamic truncation
                offset = hmac_hash[-1] & 0x0F
                code = struct.unpack('>I', hmac_hash[offset:offset+4])[0] & 0x7FFFFFFF
                return str(code % 1000000).zfill(6)

            # Check form exists
            form_check = evaluate('''
                (function() {
                    let email = document.querySelector('#email');
                    let pass = document.querySelector('#pass');
                    return email && pass ? 'OK' : 'NO_FORM';
                })()
            ''')

            if form_check != 'OK':
                ws.close()
                api.close_browser(uuid)
                return False, 'NO_FORM'

            # Type email/phone (giả lập gõ từng ký tự)
            self.after(0, lambda: self._log(f"  Typing email..."))
            type_text(account["fb_id"], "#email")

            # Pause như người chuyển field
            time.sleep(random.uniform(0.3, 0.7))

            # Type password
            self.after(0, lambda: self._log(f"  Typing password..."))
            type_text(account["password"], "#pass")

            # Pause trước khi click
            time.sleep(random.uniform(0.5, 1.0))

            # Click login button
            login_result = evaluate('''
                (function() {
                    let loginBtn = document.querySelector('#loginbutton') ||
                                   document.querySelector('button[name="login"]') ||
                                   document.querySelector('button[type="submit"]');
                    if (loginBtn) {
                        loginBtn.click();
                        return 'CLICKED';
                    }
                    // Fallback: submit form
                    let form = document.querySelector('#email')?.closest('form');
                    if (form) {
                        form.submit();
                        return 'SUBMITTED';
                    }
                    return 'NO_BTN';
                })()
            ''')

            self.after(0, lambda r=login_result: self._log(f"  Login form: {r}"))

            if login_result not in ['CLICKED', 'SUBMITTED']:
                ws.close()
                api.close_browser(uuid)
                return False, 'NO_FORM'

            # Wait for login result - đợi page load xong
            time.sleep(2)

            # Đợi page load hoàn tất (max 10s)
            for _ in range(10):
                ready_state = evaluate('document.readyState')
                if ready_state == 'complete':
                    break
                time.sleep(1)

            time.sleep(1 + random.uniform(0.5, 1.5))

            # Check login result
            js_check = '''
                (function() {
                    let url = window.location.href;
                    let pageText = (document.body.innerText || '').toLowerCase();

                    // Check for checkpoint/locked/verify
                    if (url.includes('checkpoint') || url.includes('locked') ||
                        url.includes('verify') || url.includes('confirm')) {
                        return 'LOCKED';
                    }

                    // Check for 2FA
                    if (url.includes('two_step') || url.includes('code_generator') ||
                        document.querySelector('input[name="approvals_code"]')) {
                        return '2FA';
                    }

                    // Check for wrong password - nhiều cách FB hiển thị
                    // 1. Check error text trực tiếp
                    let wrongPassPhrases = [
                        'password that you',
                        'password you entered',
                        'incorrect password',
                        'wrong password',
                        'sai mật khẩu',
                        'mật khẩu không đúng',
                        'mật khẩu bạn nhập',
                        'forgotten password'
                    ];
                    for (let phrase of wrongPassPhrases) {
                        if (pageText.includes(phrase)) return 'WRONG_PASS';
                    }

                    // 2. Check error elements
                    let errorEl = document.querySelector('._9ay7, [data-testid="royal_login_form"] + div, .login_error_box');
                    if (errorEl && errorEl.innerText.toLowerCase().includes('password')) {
                        return 'WRONG_PASS';
                    }

                    // Check for disabled/banned account
                    let diePhrases = [
                        'vô hiệu hóa tài khoản',
                        'đã vô hiệu hóa',
                        'bị vô hiệu hóa',
                        'disabled',
                        'suspended',
                        'đã bị khóa',
                        'account has been disabled',
                        'tài khoản của bạn đã bị'
                    ];
                    for (let phrase of diePhrases) {
                        if (pageText.includes(phrase)) return 'DIE';
                    }

                    // Check if logged in
                    let isLoggedIn = document.querySelector('[aria-label*="Account"]') ||
                                     document.querySelector('[aria-label*="Tài khoản"]') ||
                                     document.querySelector('[aria-label*="Messenger"]') ||
                                     document.querySelector('div[role="navigation"]') ||
                                     document.querySelector('[aria-label="Facebook"]');

                    if (isLoggedIn || url === 'https://www.facebook.com/' ||
                        url.includes('facebook.com/?sk=') || url.includes('facebook.com/home')) {
                        return 'LIVE';
                    }

                    // Still on login page after submit = likely wrong credentials
                    if (url.includes('/login') && document.querySelector('#pass')) {
                        // Check if there's any visible error indicator
                        let hasError = document.querySelector('[role="alert"], .uiBoxRed, ._9ay7');
                        if (hasError) return 'WRONG_PASS';
                        return 'FAILED';
                    }

                    return 'UNKNOWN:' + url.substring(0, 50);
                })()
            '''

            # Retry check status (đợi nếu UNKNOWN)
            status = 'UNKNOWN'
            status_clean = 'UNKNOWN'
            for attempt in range(3):
                status = evaluate(js_check) or 'UNKNOWN'
                status_clean = status.split(':')[0] if ':' in str(status) else status

                # Nếu có kết quả rõ ràng thì dừng
                if status_clean in ['LIVE', 'DIE', 'WRONG_PASS', 'LOCKED', '2FA']:
                    break

                # Chờ thêm nếu UNKNOWN/FAILED
                if attempt < 2:
                    time.sleep(2)

            self.after(0, lambda s=status: self._log(f"  Login status: {s}"))

            # Xử lý 2FA nếu có secret key
            if status_clean == '2FA' and account.get('totp_secret'):
                try:
                    totp_code = generate_totp(account['totp_secret'])
                    self.after(0, lambda c=totp_code: self._log(f"  2FA code: {c}"))

                    time.sleep(1)

                    # Bước 0: Kiểm tra có input 2FA không, nếu không thì click "Thử cách khác"
                    has_input = evaluate('''
                        (function() {
                            let input = document.querySelector('input[id^="_r_"]') ||
                                       document.querySelector('input[autocomplete="off"][type="text"]') ||
                                       document.querySelector('input[name="approvals_code"]');
                            return input ? 'YES' : 'NO';
                        })()
                    ''')
                    self.after(0, lambda h=has_input: self._log(f"  2FA input found: {h}"))

                    if has_input != 'YES':
                        # Click "Thử cách khác" / "Try another way"
                        self.after(0, lambda: self._log(f"  Clicking 'Thử cách khác'..."))
                        evaluate('''
                            (function() {
                                let btns = document.querySelectorAll('div[role="button"], button, a, span');
                                for (let btn of btns) {
                                    let text = (btn.innerText || btn.textContent || '').trim();
                                    if (text.includes('Thử cách khác') || text.includes('Try another') ||
                                        text.includes('other way') || text.includes('cách khác')) {
                                        btn.click();
                                        return 'clicked_try_another';
                                    }
                                }
                                return 'no_try_another';
                            })()
                        ''')
                        time.sleep(2)

                    # Bước 1: Click vào "Ứng dụng xác thực" nếu đang ở màn chọn phương thức
                    evaluate('''
                        (function() {
                            // Method 1: Click radio input value="1" (authentication app)
                            let radio = document.querySelector('input[type="radio"][value="1"]');
                            if (radio) {
                                // Click parent container to trigger selection
                                let container = radio;
                                for (let i = 0; i < 8; i++) {
                                    container = container.parentElement;
                                    if (!container) break;
                                }
                                if (container) container.click();
                                radio.click();
                                radio.checked = true;
                                radio.dispatchEvent(new Event('change', {bubbles: true}));
                                return 'clicked_radio_v1';
                            }

                            // Method 2: Find by text "Ứng dụng xác thực"
                            let allEls = document.querySelectorAll('div, span');
                            for (let el of allEls) {
                                let text = (el.innerText || el.textContent || '').trim();
                                if (text === 'Ứng dụng xác thực' || text === 'Authentication app') {
                                    // Walk up to find clickable parent with radio
                                    let parent = el;
                                    for (let i = 0; i < 10; i++) {
                                        parent = parent.parentElement;
                                        if (!parent) break;
                                        let radio = parent.querySelector('input[type="radio"]');
                                        if (radio) {
                                            parent.click();
                                            radio.click();
                                            radio.checked = true;
                                            return 'clicked_radio_v2';
                                        }
                                    }
                                }
                            }
                            return 'no_auth_option';
                        })()
                    ''')
                    time.sleep(1.5)

                    # Bước 2: Click nút "Tiếp tục" sau khi chọn phương thức
                    evaluate('''
                        (function() {
                            let btns = document.querySelectorAll('div[role="button"], button');
                            for (let btn of btns) {
                                let text = (btn.innerText || '').trim();
                                // Tránh click "Thử cách khác"
                                if (text.includes('Thử cách khác') || text.includes('Try another')) continue;
                                if (text.includes('Tiếp tục') || text.includes('Continue') || text.includes('Next')) {
                                    btn.click();
                                    return 'clicked_continue';
                                }
                            }
                            return 'no_continue';
                        })()
                    ''')
                    time.sleep(2)

                    # Bước 3: Focus vào input 2FA (selector mới cho FB React)
                    evaluate('''
                        (function() {
                            let input = document.querySelector('input[id^="_r_"]') ||
                                       document.querySelector('input[autocomplete="off"][type="text"]') ||
                                       document.querySelector('input[name="approvals_code"]');
                            if (input) {
                                input.focus();
                                // Dùng native setter để bypass React
                                const nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                                nativeInputValueSetter.call(input, '');
                                input.dispatchEvent(new Event('input', {bubbles: true}));
                            }
                        })()
                    ''')
                    time.sleep(0.3)

                    # Gõ từng số của 2FA code
                    for digit in totp_code:
                        send_cmd("Input.insertText", {"text": digit})
                        time.sleep(random.uniform(0.15, 0.35))

                    time.sleep(0.8)

                    # Submit 2FA form - tìm theo text
                    evaluate('''
                        (function() {
                            let btns = document.querySelectorAll('div[role="button"], button');
                            for (let btn of btns) {
                                let text = (btn.innerText || '').toLowerCase();
                                if (text.includes('tiếp tục') || text.includes('continue') ||
                                    text.includes('gửi') || text.includes('submit')) {
                                    btn.click();
                                    return 'clicked';
                                }
                            }
                            // Fallback: aria-label
                            let btn = document.querySelector('div[aria-label*="Tiếp"]') ||
                                     document.querySelector('div[aria-label*="Continue"]') ||
                                     document.querySelector('button[type="submit"]');
                            if (btn) { btn.click(); return 'clicked_fallback'; }
                            return 'no_btn';
                        })()
                    ''')

                    # Đợi và check lại
                    time.sleep(4)
                    status = evaluate(js_check) or 'UNKNOWN'
                    status_clean = status.split(':')[0] if ':' in str(status) else status
                    self.after(0, lambda s=status: self._log(f"  After 2FA: {s}"))

                except Exception as e:
                    self.after(0, lambda err=str(e): self._log(f"  2FA error: {err}"))

            ws.close()

            # Đóng browser TRƯỚC khi xóa profile
            if status_clean != 'LIVE':
                api.close_browser(uuid)
                time.sleep(0.5)  # Đợi browser đóng xong

            # Xóa profile nếu login thất bại (không phải LIVE) và option được bật
            if status_clean != 'LIVE' and self.delete_bad_var.get():
                try:
                    self.after(0, lambda s=status_clean: self._log(f"  🗑️ Xóa profile ({s})..."))
                    # Thử xóa local trước, nếu fail thì thử remote
                    delete_result = api.delete_profiles([uuid], is_local=True)
                    self.after(0, lambda r=str(delete_result): self._log(f"  Delete result: {r}"))
                    # Nếu không thành công, thử remote
                    if not delete_result.get('status') == 'successfully':
                        delete_result = api.delete_profiles([uuid], is_local=False)
                        self.after(0, lambda r=str(delete_result): self._log(f"  Delete remote: {r}"))
                except Exception as e:
                    self.after(0, lambda err=str(e): self._log(f"  Delete error: {err}"))

            return status_clean == 'LIVE', status_clean

        except Exception as e:
            print(f"Login error: {e}")
            if ws:
                try:
                    ws.close()
                except:
                    pass
            # Đóng browser trước
            api.close_browser(uuid)
            time.sleep(0.5)
            # Xóa profile khi có lỗi
            if self.delete_bad_var.get():
                try:
                    self.after(0, lambda: self._log(f"  🗑️ Xóa profile (ERROR)..."))
                    api.delete_profiles([uuid], is_local=False)
                except:
                    pass
            return False, 'ERROR'

    def _stop_login(self):
        """Dừng login"""
        self._stop_requested = True
        self._log("⏹️ Đang dừng...")

    def _on_login_complete(self):
        """Khi login xong"""
        self._is_running = False
        self._render_accounts()
        self._apply_filter()
        self._log("✅ Đã hoàn thành login")

    def _log(self, message: str):
        """Ghi log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")

        if self.status_callback:
            self.status_callback(message)
